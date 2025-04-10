import sys
import os
import tkinter as tk
from tkinter import simpledialog, messagebox
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from .layer_selector import selecionar_layers
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from ui.talhoes_parser import extrair_talhoes_por_proximidade,extrair_legenda_layers
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt
from PySide6.QtWidgets import QApplication
from ui.excel_viewer import PDFViewer
import subprocess
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage
from ui.imagem_utils import redimensionar_imagem, gerar_imagem_centrada, inserir_imagem
from openpyxl.worksheet.page import PageMargins
from ui.layout_dialog import ExtendedLayoutInfoDialog as LayoutInfoDialog
from PySide6.QtWidgets import QApplication, QMessageBox

MAX_DESENHISTA = 60  # Limite máximo para o nome do DESENHISTA

def set_cell_value(ws, cell_coord, value):
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            anchor = merged_range.start_cell.coordinate
            ws[anchor].value = value
            return
    ws[cell_coord].value = value

def converter_excel_para_pdf_com_libreoffice(excel_path):
    """
    Converte um arquivo .xlsx para .pdf usando LibreOffice Portable.
    """
    try:
        # Caminho do LibreOffice dentro do seu projeto
        libreoffice_path = os.path.abspath(
            r"LibreOfficePortable\App\libreoffice\program\soffice.exe"
        )
        print(f"📌 Caminho do LibreOffice: {libreoffice_path}")
        print(f"📌 Excel de entrada: {excel_path}")

        if not os.path.exists(libreoffice_path):
            raise FileNotFoundError("❌ LibreOffice Portable não encontrado no caminho especificado.")

        excel_path = os.path.abspath(excel_path)
        output_dir = os.path.dirname(excel_path)

        print(f"📌 Diretório de saída: {output_dir}")
        
        # Montar o comando
        command = [
            libreoffice_path,
            "--headless",
            "--convert-to", "pdf",
            "--outdir", output_dir,
            excel_path
        ]
        print("📤 Comando que será executado:")
        print(" ".join(command))

        # Executar o comando
        subprocess.run(command, check=True)

        pdf_path = excel_path.replace(".xlsx", ".pdf")
        print(f"📌 Esperando PDF em: {pdf_path}")
        print(f"📌 PDF existe? {os.path.exists(pdf_path)}")

        if os.path.exists(pdf_path):
            print(f"✅ PDF gerado com sucesso: {pdf_path}")
            return pdf_path
        else:
            print("❌ PDF não foi gerado.")
            return None

    except Exception as e:
        print(f"❌ Erro ao converter Excel para PDF: {e}")
        return None

    except Exception as e:
        print(f"❌ Erro ao converter Excel para PDF: {e}")
        return None
    except Exception as e:
        print(f"❌ Erro ao converter Excel para PDF: {e}")
        return None
    
def preparar_planilha_para_pdf(wb, escalas_por_aba=None, escala_padrao=75, print_areas=None):
    """
    Configura as abas para exportação em PDF centralizado.
    """
    if escalas_por_aba is None:
        escalas_por_aba = {}
    if print_areas is None:
        print_areas = {}

    for ws in wb.worksheets:
        escala = escalas_por_aba.get(ws.title, escala_padrao)
        area = print_areas.get(ws.title)

        ws.page_margins = PageMargins(
            left=0.3, right=0.3, top=0.4, bottom=0.4,
            header=0.0, footer=0.0
        )
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.scale = escala
        ws.page_setup.fitToWidth = False
        ws.page_setup.fitToHeight = False

        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True

        if area:
            ws.print_area = area

def adicionar_legenda_layers(ws, legenda_layers, start_row=1, start_col=1):
    """
    Gera uma legenda baseada no dicionário {layer_name: {"color": (r, g, b)}}.
    Cada linha terá:
      - 1 célula com cor de fundo (PatternFill)
      - 1 célula com o nome do layer
    """
    # Exemplo de título "PROJETO DE SISTEMATIZAÇÃO"
    titulo = "PROJETO DE SISTEMATIZAÇÃO"
    titulo_cell = ws.cell(row=start_row, column=start_col, value=titulo)
    titulo_cell.font = Font(bold=True, size=14)
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 2
    )
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Deixe uma linha em branco após o título
    row = start_row + 2

    for layer_name, info in legenda_layers.items():
        color_floats = info["color"]  # Ex.: (0.0, 1.0, 0.0) para verde
        r_float, g_float, b_float = color_floats
        # Converter floats [0..1] em RGB hex (ex.: "FF00FF")
        r = int(r_float * 255)
        g = int(g_float * 255)
        b = int(b_float * 255)
        color_hex = f"{r:02X}{g:02X}{b:02X}"

        # Celula colorida
        color_cell = ws.cell(row=row, column=start_col)
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        color_cell.fill = fill

        # Celula com o nome do layer ao lado
        name_cell = ws.cell(row=row, column=start_col + 1, value=layer_name)
        name_cell.alignment = Alignment(horizontal="left", vertical="center")

        row += 1

def redimensionar_imagem(imagem_path, largura, altura):
    try:
        with Image.open(imagem_path) as img:
            resized_img = img.resize((largura, altura), Image.LANCZOS)
            resized_img.save(imagem_path)
            print("✅ Imagem redimensionada para:", resized_img.size)
    except Exception as e:
        print(f"❌ Erro ao redimensionar imagem: {e}")

def limpar_colunas_fora_do_layout(ws, ultima_coluna_valida="K"):
    col_idx = openpyxl.utils.column_index_from_string(ultima_coluna_valida)
    for i in range(col_idx + 1, 100):  # limpa colunas de L até CV
        col = get_column_letter(i)
        if col in ws.column_dimensions:
            del ws.column_dimensions[col]

def limpar_linhas_fora_do_layout(ws, ultima_linha_valida=33):
    for i in range(ultima_linha_valida + 1, 200):  # limpa linhas 34 em diante
        if i in ws.row_dimensions:
            del ws.row_dimensions[i]

def adicionar_tabela_comprimentos_custom(ws, layer_data, start_row=1, start_col=1):
    # --------------------------
    #   Configurações de estilo
    # --------------------------
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=False)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=False)
    
    # --------------------------
    #   Montar título mesclado
    # --------------------------
    title_cell = ws.cell(row=start_row, column=start_col, value="COMPRIMENTOS POR LAYER")
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 3
    )
    
    # --------------------------
    #   Cabeçalho
    # --------------------------
    headers = ["NOME DO LAYER", "QTD", "TOTAL (m)", "MÉDIA (m)"]
    header_row = start_row + 1
    for i, header_text in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i, value=header_text)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # --------------------------
    #   Inserir dados
    # --------------------------
    data_start_row = header_row + 1
    current_row = data_start_row
    
    for layer, data in layer_data.items():
        qtd = data.get("qtd", 0)
        total_m = data.get("total", 0.0)
        media_m = total_m / qtd if qtd > 0 else 0.0
        
        # Coluna 1: NOME DO LAYER
        cell = ws.cell(row=current_row, column=start_col, value=layer)
        cell.font = cell_font
        # Aqui, usamos left_alignment com wrap_text=False para manter o tamanho fixo
        cell.alignment = left_alignment
        cell.border = thin_border
        
        # Coluna 2: QTD
        cell = ws.cell(row=current_row, column=start_col + 1, value=qtd)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Coluna 3: TOTAL (m)
        cell = ws.cell(row=current_row, column=start_col + 2, value=round(total_m, 2))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Coluna 4: MÉDIA (m)
        cell = ws.cell(row=current_row, column=start_col + 3, value=round(media_m, 2))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        current_row += 1

def parse_talhao_layer_name(layer_name):
    """
    Recebe algo como '06.11.14' e retorna ('06', 11.14).
    Se não houver ponto ou não for possível converter a área, 
    retorna (layer_name, 0.0).
    """
    parts = layer_name.split('.', 1)  # Divide em 2 partes no primeiro ponto
    if len(parts) == 2:
        numero_str, area_str = parts
        numero_str = numero_str.strip()
        try:
            area_ha = float(area_str)
        except ValueError:
            area_ha = 0.0
    else:
        # Se não houver ponto, ou não der para converter, 
        # assume o layer_name inteiro como número e área = 0
        numero_str = layer_name
        area_ha = 0.0

    return numero_str, area_ha

def adicionar_tabela_talhoes_custom(ws, talhoes_dict, start_row=1, start_col=1):
    """
    Cria uma tabela "TALHÕES" sem a coluna de %,
    exibindo:
      - Número
      - Área (ha)
      - Área (alq)*

    E uma linha TOTAL em vermelho.
    """
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Título
    ws.cell(row=start_row, column=start_col, value="TALHÕES").font = title_font
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 2
    )
    ws.cell(row=start_row, column=start_col).alignment = center_alignment

    # Cabeçalho
    headers = ["Número", "Área (ha)", "Área (alq)*"]
    for i, header_text in enumerate(headers):
        c = ws.cell(row=start_row+1, column=start_col + i, value=header_text)
        c.font = header_font
        c.alignment = center_alignment
        c.border = thin_border

    # Inserir dados
    total_ha = 0.0
    total_alq = 0.0
    row_data_start = start_row + 2

    current_row = row_data_start
    for numero, area_ha in talhoes_dict.items():
        area_alq = area_ha / 2.42  # Ajuste se quiser outro fator

        # Número
        c = ws.cell(row=current_row, column=start_col, value=numero)
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        # Área (ha)
        c = ws.cell(row=current_row, column=start_col+1, value=round(area_ha, 2))
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        # Área (alq)*
        c = ws.cell(row=current_row, column=start_col+2, value=round(area_alq, 2))
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        total_ha += area_ha
        total_alq += area_alq
        current_row += 1

    # Linha TOTAL
    c = ws.cell(row=current_row, column=start_col, value="TOTAL")
    c.font = Font(bold=True, color="FF0000")  # vermelho
    c.alignment = center_alignment
    c.border = thin_border

    c = ws.cell(row=current_row, column=start_col+1, value=round(total_ha, 2))
    c.font = Font(bold=True)
    c.alignment = center_alignment
    c.border = thin_border

    c = ws.cell(row=current_row, column=start_col+2, value=round(total_alq, 2))
    c.font = Font(bold=True)
    c.alignment = center_alignment
    c.border = thin_border

    # Observação "*Alqueires Paulistas"
    ws.cell(row=current_row+1, column=start_col+2, value="*Alqueires Paulistas").alignment = Alignment(horizontal="right")

    # Ajuste de largura (remova se quiser manter o template)
    ws.column_dimensions[get_column_letter(start_col)].width = 10
    ws.column_dimensions[get_column_letter(start_col+1)].width = 12
    ws.column_dimensions[get_column_letter(start_col+2)].width = 12

def gerar_layout_final(dxf_file_path, layer_data, talhoes_dict, legenda_layers, dados):
    # Aqui você usa 'dados' diretamente, sem abrir outro diálogo.
    print("Dados recebidos:", dados)

    def gerar_nome_excel(dxf_file_path, versao_anterior=None):
        nome_dxf = os.path.splitext(os.path.basename(dxf_file_path))[0]
        if versao_anterior is None:
            versao = 0.1
        else:
            try:
                versao = round(float(versao_anterior) + 0.1, 1)
            except ValueError:
                versao = 0.1
        return f"{nome_dxf}_V{versao}.xlsx"

    def centralizar_imagem_na_planilha(ws, imagem_path, cell_coord="E20"):
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        print("centralizar_imagem_na_planilha chamada com imagem_path:", imagem_path)
        if not os.path.exists(imagem_path):
            print("❌ Imagem do mapa não foi encontrada.")
            return
        try:
            img = XLImage(imagem_path)
            cell = ws[cell_coord]
            col_letter = get_column_letter(cell.column)
            row_num = cell.row
            img.anchor = f"{col_letter}{row_num}"
            ws.add_image(img)
            print(f"✅ Imagem inserida na planilha na célula {cell_coord}")
        except Exception as e:
            print(f"❌ Erro ao inserir imagem na planilha: {e}")

    def resource_path(relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    template_file = resource_path('resources/excel/Planilha_template.xlsx')

    # Usa o diretório definido pelo usuário (via interface), ou fallback para 'output'
    output_dir = dados.get("out_dir", os.path.join(os.path.dirname(__file__), '..', 'output'))
    os.makedirs(output_dir, exist_ok=True)

    output_file = os.path.join(output_dir, gerar_nome_excel(dxf_file_path))

    wb = openpyxl.load_workbook(template_file)
    if "Pagina1" not in wb.sheetnames or "Pagina2" not in wb.sheetnames:
        print("❌ As abas 'Pagina1' ou 'Pagina2' não foram encontradas no template.")
        return

    ws_pagina1 = wb["Pagina1"]
    ws_pagina2 = wb["Pagina2"]

    # Supondo que as funções abaixo já estão definidas em seu projeto:
    # limpar_colunas_fora_do_layout(ws, ultima_coluna_valida)
    # limpar_linhas_fora_do_layout(ws, ultima_linha_valida)
    # preparar_planilha_para_pdf(wb, escalas_por_aba, print_areas)
    # adicionar_legenda_layers(ws, legenda_layers, start_row, start_col)
    # adicionar_tabela_comprimentos_custom(ws, layer_data, start_row, start_col)
    # adicionar_tabela_talhoes_custom(ws, talhoes_dict, start_row, start_col)
    # set_cell_value(ws, cell_coord, value)

    try:
        from ui.layout_generator import (
            limpar_colunas_fora_do_layout, limpar_linhas_fora_do_layout,
            preparar_planilha_para_pdf, adicionar_legenda_layers,
            adicionar_tabela_comprimentos_custom, adicionar_tabela_talhoes_custom, set_cell_value
        )
    except ImportError as e:
        print("Erro ao importar funções auxiliares:", e)
        return

    limpar_colunas_fora_do_layout(ws_pagina1, "K")
    limpar_linhas_fora_do_layout(ws_pagina1, 33)
    limpar_colunas_fora_do_layout(ws_pagina2, "J")
    limpar_linhas_fora_do_layout(ws_pagina2, 33)

    preparar_planilha_para_pdf(
        wb,
        escalas_por_aba={"Pagina1": 75, "Pagina2": 85},
        print_areas={"Pagina1": "A1:K33", "Pagina2": "A1:J33"}
    )

    ws_pagina1.merge_cells("H33:I33")
    ws_pagina2.merge_cells("F33:J33")

    # Logo da Cevasa
    img_cevasa_path = resource_path("resources/images/logo.png")
    redimensionar_imagem(img_cevasa_path, 95, 40)
    img_cevasa = XLImage(img_cevasa_path)
    img_cevasa.anchor = "A32"
    ws_pagina2.add_image(img_cevasa)
    ws_pagina1.column_dimensions["K"].width = 36

    # Rosa dos ventos
    img_rosa_path_1 = resource_path("resources/images/rosa_dos_ventos.png")
    redimensionar_imagem(img_rosa_path_1, 110, 110)
    img_rosa_path_2 = resource_path("resources/images/rosa_dos_ventos.png")
    redimensionar_imagem(img_rosa_path_2, 100, 90)

    ws_pagina1.merge_cells("K28:K31")
    ws_pagina2.merge_cells("I28:J31")

    img_final_rosa_1 = os.path.join("output", "rosa_dos_ventos_pagina1.png")
    img_final_rosa_2 = os.path.join("output", "rosa_dos_ventos_pagina2.png")
    gerar_imagem_centrada(img_rosa_path_1, 252, 110, img_final_rosa_1)
    inserir_imagem(ws_pagina1, img_final_rosa_1, "K28")
    gerar_imagem_centrada(img_rosa_path_2, 170, 90, img_final_rosa_2)
    inserir_imagem(ws_pagina2, img_final_rosa_2, "I28")
    ws_pagina1.column_dimensions["K"].width = 36

    img_cevasa = XLImage(resource_path("resources/images/logo.png"))
    img_cevasa.anchor = "A32"
    ws_pagina1.add_image(img_cevasa)

    # Preencher informações na planilha
    set_cell_value(ws_pagina1, "I28", dados['parc'])
    set_cell_value(ws_pagina1, "J29", dados['data_atual'])
    set_cell_value(ws_pagina1, "I30", dados['distancia'])
    set_cell_value(ws_pagina1, "I31", dados['area_cana'])
    set_cell_value(ws_pagina1, "J31", dados['nova_versao'])
    set_cell_value(ws_pagina1, "I29", dados['escala'])
    set_cell_value(ws_pagina1, "B33", dados['propriedade'])
    set_cell_value(ws_pagina1, "E33", dados['mun_est'])
    set_cell_value(ws_pagina1, "H33", dados['desenhista'])

    set_cell_value(ws_pagina2, "G28", dados['parc'])
    set_cell_value(ws_pagina2, "H29", dados['data_atual'])
    set_cell_value(ws_pagina2, "G30", dados['distancia'])
    set_cell_value(ws_pagina2, "G31", dados['area_cana'])
    set_cell_value(ws_pagina2, "H31", dados['nova_versao'])
    set_cell_value(ws_pagina2, "G29", dados['escala'])
    set_cell_value(ws_pagina2, "B33", dados['propriedade'])
    set_cell_value(ws_pagina2, "C33", dados['mun_est'])
    set_cell_value(ws_pagina2, "F33", dados['desenhista'])

    adicionar_tabela_comprimentos_custom(ws_pagina2, layer_data, start_row=2, start_col=2)
    adicionar_tabela_talhoes_custom(ws_pagina2, talhoes_dict, start_row=2, start_col=7)
    adicionar_legenda_layers(ws_pagina1, legenda_layers, start_row=1, start_col=9)

    image_path = os.path.join("output", "mapa.png")
    if os.path.exists(image_path):
        try:
            redimensionar_imagem(image_path, 800, 575)
            centralizar_imagem_na_planilha(ws_pagina1, image_path, cell_coord="A02")
            print("✅ Imagem 'mapa.png' adicionada na aba 'Pagina1'.")
        except Exception as e:
            print(f"❌ Erro ao inserir imagem 'mapa.png': {e}")
    else:
        print("❌ Imagem do mapa não foi encontrada no caminho:", image_path)

    wb.save(output_file)
    print(f"✅ Planilha salva como '{output_file}'.")

    pdf_path = converter_excel_para_pdf_com_libreoffice(output_file)
    if pdf_path:
        QMessageBox.information(None, "PDF Gerado", f"PDF gerado com sucesso:\n{pdf_path}")
    else:
        QMessageBox.warning(None, "Erro", "Não foi possível gerar o PDF da planilha final.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    # Para teste: substitua os parâmetros pelos reais ou fictícios
    gerar_layout_final("exemplo.dxf",
                       {"Layer1": {"qtd": 5, "total": 100.0}},
                       {"01": 5.0},
                       {"Layer1": {"color": (0.0, 1.0, 0.0)}})
    sys.exit(app.exec())