import sys
import os
import tkinter as tk
from tkinter import simpledialog, messagebox
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from .layer_selector import selecionar_layers
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from src.talhoes_parser import extrair_talhoes_por_proximidade,extrair_legenda_layers
import win32com.client as win32
from openpyxl.styles import Font, Alignment
import matplotlib.pyplot as plt

MAX_DESENHISTA = 24  # Limite máximo para o nome do DESENHISTA

def ask_limited_string(prompt, limit=MAX_DESENHISTA):
    def validate_input(P):
        return len(P) <= limit

    dlg = tk.Toplevel()
    dlg.title(prompt)
    dlg.resizable(False, False)
    dlg.geometry("350x150+400+250")

    tk.Label(dlg, text=f"{prompt} (máx {limit} caracteres)").pack(padx=10, pady=5)

    var = tk.StringVar()
    entry = tk.Entry(dlg, textvariable=var, validate='key',
                     validatecommand=(dlg.register(validate_input), '%P'))
    entry.pack(padx=10, pady=5)
    entry.focus()

    result = None

    def on_ok():
        nonlocal result
        result = var.get().strip().upper()
        if not result:
            messagebox.showerror("Erro", "Campo obrigatório.", parent=dlg)
        else:
            dlg.destroy()

    tk.Button(dlg, text="OK", command=on_ok).pack(pady=10)
    dlg.grab_set()
    dlg.wait_window()

    return result
def excel_to_pdf(excel_path, pdf_path):
    """
    Converte um arquivo Excel para PDF utilizando o Microsoft Excel via COM.
    Requer que o Excel esteja instalado no sistema (funciona apenas no Windows).
    """
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        # Abre o workbook
        workbook = excel.Workbooks.Open(os.path.abspath(excel_path))
        # Exporta como PDF (0 indica PDF)
        workbook.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        workbook.Close(False)
        print(f"✅ Planilha convertida para PDF: {pdf_path}")
    except Exception as e:
        print(f"❌ Erro ao converter Excel para PDF: {e}")
    finally:
        excel.Quit()

def set_cell_value(ws, cell_coord, value):
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            anchor = merged_range.start_cell.coordinate
            ws[anchor].value = value
            return
    ws[cell_coord].value = value

def adicionar_legenda_layers(ws, legenda_layers, start_row=1, start_col=1):
    """
    Gera uma legenda baseada no dicionário {layer_name: {"color": (r, g, b)}}.
    Cada linha terá:
      - 1 célula com cor de fundo (PatternFill)
      - 1 célula com o nome do layer
    """
    # Exemplo de título "PROJETO DE SISTEMATIZAÇÃO"
    titulo = "PROJETO DE SISTEMATIZAÇÃO"
    titulo_cell = ws.cell(row=start_row, column=start_col, value=titulo)
    titulo_cell.font = Font(bold=True, size=14)
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 2
    )
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Deixe uma linha em branco após o título
    row = start_row + 2

    for layer_name, info in legenda_layers.items():
        color_floats = info["color"]  # Ex.: (0.0, 1.0, 0.0) para verde
        r_float, g_float, b_float = color_floats
        # Converter floats [0..1] em RGB hex (ex.: "FF00FF")
        r = int(r_float * 255)
        g = int(g_float * 255)
        b = int(b_float * 255)
        color_hex = f"{r:02X}{g:02X}{b:02X}"

        # Celula colorida
        color_cell = ws.cell(row=row, column=start_col)
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        color_cell.fill = fill

        # Celula com o nome do layer ao lado
        name_cell = ws.cell(row=row, column=start_col + 1, value=layer_name)
        name_cell.alignment = Alignment(horizontal="left", vertical="center")

        row += 1

def redimensionar_imagem(imagem_path, largura, altura):
    try:
        with Image.open(imagem_path) as img:
            resized_img = img.resize((largura, altura), Image.LANCZOS)
            resized_img.save(imagem_path)
            print("✅ Imagem redimensionada para:", resized_img.size)
    except Exception as e:
        print(f"❌ Erro ao redimensionar imagem: {e}")

def adicionar_tabela_comprimentos_custom(ws, layer_data, start_row=1, start_col=1):
    # --------------------------
    #   Configurações de estilo
    # --------------------------
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=False)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=False)
    
    # --------------------------
    #   Montar título mesclado
    # --------------------------
    title_cell = ws.cell(row=start_row, column=start_col, value="COMPRIMENTOS POR LAYER")
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 3
    )
    
    # --------------------------
    #   Cabeçalho
    # --------------------------
    headers = ["NOME DO LAYER", "QTD", "TOTAL (m)", "MÉDIA (m)"]
    header_row = start_row + 1
    for i, header_text in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + i, value=header_text)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # --------------------------
    #   Inserir dados
    # --------------------------
    data_start_row = header_row + 1
    current_row = data_start_row
    
    for layer, data in layer_data.items():
        qtd = data.get("qtd", 0)
        total_m = data.get("total", 0.0)
        media_m = total_m / qtd if qtd > 0 else 0.0
        
        # Coluna 1: NOME DO LAYER
        cell = ws.cell(row=current_row, column=start_col, value=layer)
        cell.font = cell_font
        # Aqui, usamos left_alignment com wrap_text=False para manter o tamanho fixo
        cell.alignment = left_alignment
        cell.border = thin_border
        
        # Coluna 2: QTD
        cell = ws.cell(row=current_row, column=start_col + 1, value=qtd)
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Coluna 3: TOTAL (m)
        cell = ws.cell(row=current_row, column=start_col + 2, value=round(total_m, 2))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        # Coluna 4: MÉDIA (m)
        cell = ws.cell(row=current_row, column=start_col + 3, value=round(media_m, 2))
        cell.font = cell_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        current_row += 1

def parse_talhao_layer_name(layer_name):
    """
    Recebe algo como '06.11.14' e retorna ('06', 11.14).
    Se não houver ponto ou não for possível converter a área, 
    retorna (layer_name, 0.0).
    """
    parts = layer_name.split('.', 1)  # Divide em 2 partes no primeiro ponto
    if len(parts) == 2:
        numero_str, area_str = parts
        numero_str = numero_str.strip()
        try:
            area_ha = float(area_str)
        except ValueError:
            area_ha = 0.0
    else:
        # Se não houver ponto, ou não der para converter, 
        # assume o layer_name inteiro como número e área = 0
        numero_str = layer_name
        area_ha = 0.0

    return numero_str, area_ha

def adicionar_tabela_talhoes_custom(ws, talhoes_dict, start_row=1, start_col=1):
    """
    Cria uma tabela "TALHÕES" sem a coluna de %,
    exibindo:
      - Número
      - Área (ha)
      - Área (alq)*

    E uma linha TOTAL em vermelho.
    """
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter

    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=10)
    cell_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Título
    ws.cell(row=start_row, column=start_col, value="TALHÕES").font = title_font
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=start_row, end_column=start_col + 2
    )
    ws.cell(row=start_row, column=start_col).alignment = center_alignment

    # Cabeçalho
    headers = ["Número", "Área (ha)", "Área (alq)*"]
    for i, header_text in enumerate(headers):
        c = ws.cell(row=start_row+1, column=start_col + i, value=header_text)
        c.font = header_font
        c.alignment = center_alignment
        c.border = thin_border

    # Inserir dados
    total_ha = 0.0
    total_alq = 0.0
    row_data_start = start_row + 2

    current_row = row_data_start
    for numero, area_ha in talhoes_dict.items():
        area_alq = area_ha / 2.42  # Ajuste se quiser outro fator

        # Número
        c = ws.cell(row=current_row, column=start_col, value=numero)
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        # Área (ha)
        c = ws.cell(row=current_row, column=start_col+1, value=round(area_ha, 2))
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        # Área (alq)*
        c = ws.cell(row=current_row, column=start_col+2, value=round(area_alq, 2))
        c.font = cell_font
        c.alignment = center_alignment
        c.border = thin_border

        total_ha += area_ha
        total_alq += area_alq
        current_row += 1

    # Linha TOTAL
    c = ws.cell(row=current_row, column=start_col, value="TOTAL")
    c.font = Font(bold=True, color="FF0000")  # vermelho
    c.alignment = center_alignment
    c.border = thin_border

    c = ws.cell(row=current_row, column=start_col+1, value=round(total_ha, 2))
    c.font = Font(bold=True)
    c.alignment = center_alignment
    c.border = thin_border

    c = ws.cell(row=current_row, column=start_col+2, value=round(total_alq, 2))
    c.font = Font(bold=True)
    c.alignment = center_alignment
    c.border = thin_border

    # Observação "*Alqueires Paulistas"
    ws.cell(row=current_row+1, column=start_col+2, value="*Alqueires Paulistas").alignment = Alignment(horizontal="right")

    # Ajuste de largura (remova se quiser manter o template)
    ws.column_dimensions[get_column_letter(start_col)].width = 10
    ws.column_dimensions[get_column_letter(start_col+1)].width = 12
    ws.column_dimensions[get_column_letter(start_col+2)].width = 12

def gerar_layout_final(dxf_file_path, layer_data, talhoes_dict, legenda_layers):
    
    manager = plt.get_current_fig_manager()
    root = manager.window  # Este é o Tk principal usado pelo Matplotlib
    image_path = os.path.join("output", "mapa.png")

    def gerar_nome_excel(dxf_file_path, versao_anterior=None):
        # Extrai o nome do arquivo DXF (sem a extensão)
        nome_dxf = os.path.splitext(os.path.basename(dxf_file_path))[0]
        
        # Calcula a versão
        if versao_anterior is None:
            versao = 0.1  # Caso não tenha versão anterior
        else:
            try:
                versao = round(float(versao_anterior) + 0.1, 1)  # Incrementa a versão
            except ValueError:
                versao = 0.1  # Se a versão anterior não for válida, começa de 0.1
        
        # Cria o nome do arquivo Excel com versão
        nome_excel = f"{nome_dxf}_V{versao}.xlsx"
        
        return nome_excel

    def centralizar_imagem_na_planilha(ws, imagem_path, cell_coord="E20"):
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage

        print("centralizar_imagem_na_planilha chamada com imagem_path:", imagem_path)
        if not os.path.exists(imagem_path):
            print("❌ Imagem do mapa não encontrada.")
            return
        try:
            img = XLImage(imagem_path)
            cell = ws[cell_coord]
            col_letter = get_column_letter(cell.column)
            row_num = cell.row
            img.anchor = f"{col_letter}{row_num}"
            ws.add_image(img)
            print(f"✅ Imagem inserida na planilha na célula {cell_coord}")
        except Exception as e:
            print(f"❌ Erro ao inserir imagem na planilha: {e}")
    def abrir_tela_informacoes(parent):
        import tkinter as tk
        from tkinter import messagebox
        from datetime import datetime
        import os

        # Cria o diálogo como Toplevel usando o parent fornecido
        dialog = tk.Toplevel(parent)
        dialog.title("Preencha as informações")
        dialog.geometry("400x500")
        
        # Para garantir que o diálogo seja modal
        dialog.grab_set()
        
        # Nome do arquivo para salvar o último desenhista
        last_desenhista_file = "last_desenhista.txt"
        last_desenhista = ""

        if os.path.exists(last_desenhista_file):
            with open(last_desenhista_file, "r", encoding="utf-8") as f:
                last_desenhista = f.read().strip()

        main_frame = tk.Frame(dialog, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        def validate_desenhista(P):
            return len(P) <= MAX_DESENHISTA

        vcmd = (dialog.register(validate_desenhista), '%P')
        result = {}

        def confirmar_informacoes():
            desenhista = entry_desenhista.get().strip().upper()
            if not desenhista:
                messagebox.showerror("Erro", "Campo obrigatório.", parent=dialog)
                return
            if len(desenhista) > MAX_DESENHISTA:
                messagebox.showerror("Erro", f"O nome do Desenhista deve ter no máximo {MAX_DESENHISTA} caracteres.", parent=dialog)
                return
            with open(last_desenhista_file, "w", encoding="utf-8") as f:
                f.write(desenhista)
            
            escala = entry_escala.get().strip()
            distancia = entry_distancia.get().strip()
            area_cana = entry_area_cana.get().strip()
            prev = entry_prev.get().strip()
            mun_est = entry_mun_est.get().strip()
            parc = entry_parc.get().strip()

            try:
                prev_version = float(prev) if prev else 0.0
            except ValueError:
                prev_version = 0.0

            nova_versao = round(prev_version + 0.1, 1)
            data_atual = datetime.now().strftime("%d/%m/%Y")
            propriedade = os.path.splitext(os.path.basename(dxf_file_path))[0].upper()

            result.update({
                'desenhista': desenhista,
                'escala': escala,
                'distancia': distancia,
                'area_cana': area_cana,
                'prev_version': prev_version,
                'nova_versao': nova_versao,
                'data_atual': data_atual,
                'propriedade': propriedade,
                'mun_est': mun_est,
                'parc': parc
            })
            dialog.destroy()

        tk.Label(main_frame, text="Desenhista (máx 24 caracteres):").pack(anchor="w", pady=(0,5))
        entry_desenhista = tk.Entry(main_frame, validate='key', validatecommand=vcmd)
        entry_desenhista.pack(fill="x", pady=(0,10))
        if last_desenhista:
            entry_desenhista.insert(0, last_desenhista)

        tk.Label(main_frame, text="Escala:").pack(anchor="w", pady=(0,5))
        entry_escala = tk.Entry(main_frame)
        entry_escala.pack(fill="x", pady=(0,10))

        tk.Label(main_frame, text="Distância:").pack(anchor="w", pady=(0,5))
        entry_distancia = tk.Entry(main_frame)
        entry_distancia.pack(fill="x", pady=(0,10))

        tk.Label(main_frame, text="Área Cana (ha):").pack(anchor="w", pady=(0,5))
        entry_area_cana = tk.Entry(main_frame)
        entry_area_cana.pack(fill="x", pady=(0,10))

        tk.Label(main_frame, text="Versão Anterior:").pack(anchor="w", pady=(0,5))
        entry_prev = tk.Entry(main_frame)
        entry_prev.pack(fill="x", pady=(0,10))

        tk.Label(main_frame, text="Mun. Est. (Município e Estado):").pack(anchor="w", pady=(0,5))
        entry_mun_est = tk.Entry(main_frame)
        entry_mun_est.pack(fill="x", pady=(0,10))

        tk.Label(main_frame, text="Parc. Outorgante (opcional):").pack(anchor="w", pady=(0,5))
        entry_parc = tk.Entry(main_frame)
        entry_parc.pack(fill="x", pady=(0,10))

        tk.Button(main_frame, text="Confirmar", command=confirmar_informacoes).pack(pady=(20,0))
                
        dialog.wait_window()
        return result if result else None
    dados = abrir_tela_informacoes(root)
    if dados is None:
        print("Operação cancelada.")
        return

    # 2. Abre a janela de seleção de layers (como Toplevel do mesmo root)
    layers_selecionadas = selecionar_layers(list(layer_data.keys()), "Selecione os Layers para as Tabelas")
    if not layers_selecionadas:
        layers_selecionadas = list(layer_data.keys())
    
    # Filtra os dados conforme a seleção
    layer_data = {layer: data for layer, data in layer_data.items() if layer in layers_selecionadas}
    legenda_layers = {layer: info for layer, info in legenda_layers.items() if layer in layers_selecionadas}

    # 3. Continuação: Carregamento do template e criação da planilha final
    def resource_path(relative_path):
        try:
            base_path = sys._MEIPASS  # PyInstaller usa isso na build
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    template_file = resource_path('resources/excel/Planilha_template.xlsx')
    output_dir = os.path.join(os.path.dirname(sys.executable), 'output') if getattr(sys, 'frozen', False) else os.path.join(os.path.dirname(__file__), '..', 'output')
    os.makedirs(output_dir, exist_ok=True)

    # Gerar o nome do arquivo Excel com a versão
    output_file = os.path.join("output", gerar_nome_excel(dxf_file_path))

    wb = openpyxl.load_workbook(template_file)
    if "Pagina1" not in wb.sheetnames or "Pagina2" not in wb.sheetnames:
        print("❌ As abas 'Pagina1' ou 'Pagina2' não foram encontradas no template.")
        return
    
    ws_pagina1 = wb["Pagina1"]
    ws_pagina2 = wb["Pagina2"]

    # Adiciona a imagem do mapa
    if os.path.exists(image_path):
        try:
            redimensionar_imagem(image_path, 800, 575)
            centralizar_imagem_na_planilha(ws_pagina1, image_path, "A02")
            print("✅ Imagem 'mapa.png' adicionada na aba 'Pagina1'.")
        except Exception as e:
            print(f"❌ Erro ao inserir imagem 'mapa.png': {e}")
    else:
        print("❌ Imagem do mapa não foi encontrada no caminho:", image_path)
    # Inserir as informações na planilha
    set_cell_value(ws_pagina1, "I28", dados['desenhista'])
    set_cell_value(ws_pagina1, "J29", dados['data_atual'])
    set_cell_value(ws_pagina1, "I30", dados['distancia'])
    set_cell_value(ws_pagina1, "I31", dados['area_cana'])
    set_cell_value(ws_pagina1, "J31", dados['nova_versao'])
    set_cell_value(ws_pagina1, "I29", dados['escala'])
    set_cell_value(ws_pagina1, "B33", dados['propriedade'])
    set_cell_value(ws_pagina1, "E33", dados['mun_est'])
    set_cell_value(ws_pagina1, "H33", dados['parc'])
    
    # Inserir as tabelas na aba Pagina2
    adicionar_tabela_comprimentos_custom(ws_pagina2, layer_data, start_row=2, start_col=2)
    adicionar_tabela_talhoes_custom(ws_pagina2, talhoes_dict, start_row=2, start_col=7)
    adicionar_legenda_layers(ws_pagina1, legenda_layers, start_row=1, start_col=9)

    # Salvar a planilha com o nome gerado
    wb.save(output_file)
    print(f"✅ Planilha salva como '{output_file}'.")
    messagebox.showinfo("Sucesso", f"Planilha salva como '{output_file}'.")